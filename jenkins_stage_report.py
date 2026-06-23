#!/usr/bin/env python3
"""
jenkins_stage_report.py
--------------------------
A utility to pull Jenkins pipeline build data and report on how long a specific
stage takes across all your jobs — in this case, the Fortify SAST scan stage.

The idea is simple: instead of manually clicking through Jenkins to figure out
which pipelines ran the stage required for you and how long it took, this script does it
automatically. It hits the Jenkins API, walks every job (including nested
folders), finds builds that ran in a given date range, extracts the stage
durations, and writes everything out to a formatted Excel report.

In this case am fetching the logs for a stage called fortify scan which does a static 
code application security testing. If you want to track a different stage (e.g. "SonarQube", "Deploy", "Test"),
just change STAGE_KEYWORDS in the CONFIG section below. Everything else works
the same way.

Requirements:
    pip install requests openpyxl
"""

import time
import json
import re
from datetime import datetime, timedelta, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.series import DataPoint

import urllib3
# Suppress SSL warnings — common in internal Jenkins setups with self-signed certs.
# If your Jenkins uses a valid cert, you can remove this and set s.verify = True in make_session().
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG — update these before running
# ─────────────────────────────────────────────────────────────────────────────

JENKINS_URL  = "https://your-jenkins-server"   # No trailing slash
JENKINS_USER = "your_username"
JENKINS_TOKEN = "your_api_token"               # Jenkins API token, not your password.
                                               # Generate one at: Jenkins → Your Profile → Configure → API Token

# Date range for the report. Defaults to yesterday.
# Change these to strings like "2024-06-01" if you want a specific range.
FROM_DATE = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
TO_DATE   = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

# How many jobs to process in parallel. Increase if you have a fast network
# and a lot of jobs; lower it if you're seeing rate-limit errors from Jenkins.
MAX_WORKERS = 5

# Raw API responses get cached here so you can re-run the Excel generation
# without hammering Jenkins again. Delete the file to force a fresh fetch.
CACHE_FILE = "jenkins_raw_cache.json"

# This is the key config if you're adapting this for a different stage.
# The script checks each pipeline stage name against these keywords (case-insensitive).
# For Fortify we use "fortify scan". For SonarQube you'd use ["sonarqube", "sonar scan"], etc.
STAGE_KEYWORDS = ["fortify scan"]

# ─────────────────────────────────────────────────────────────────────────────


def get_output_filename():
    """Build the output .xlsx filename based on the date range."""
    return f"fortify_report_{FROM_DATE}_to_{TO_DATE}.xlsx"


def make_session():
    """
    Set up a requests session with Jenkins credentials and common headers.
    Using a session means we reuse the TCP connection across many API calls,
    which matters when you're hitting hundreds of jobs.
    """
    s = requests.Session()
    s.auth = (JENKINS_USER, JENKINS_TOKEN)
    s.headers.update({"Accept": "application/json"})
    s.verify = False  # Set to True or a CA bundle path if your Jenkins has a valid cert
    return s


def get_date_range_ms():
    """
    Convert FROM_DATE and TO_DATE into millisecond timestamps for comparison
    against Jenkins build timestamps (which are Unix ms).
    TO_DATE is bumped by 1 day so builds that ran on that date are included.
    """
    from_dt   = datetime.strptime(FROM_DATE, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    to_dt     = datetime.strptime(TO_DATE,   "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
    cutoff_ms = int(from_dt.timestamp() * 1000)
    end_ms    = int(to_dt.timestamp()   * 1000)
    return cutoff_ms, end_ms


def get_all_jobs(s, cutoff_ms):
    """
    Fetch every job from Jenkins, including jobs inside folders.
    Jenkins organises jobs in a tree — pipelines can be nested inside
    folders which can themselves be nested inside other folders. The 'tree'
    query parameter lets us fetch all of that in a single API call up to
    depth 7. We then flatten the tree into a plain list.

    To avoid wasting time on jobs that haven't run recently, we pre-filter:
    if a job's lastBuild timestamp is before our FROM_DATE, we skip it.
    """
    url = f"{JENKINS_URL}/api/json"
    params = {
        # This tree query fetches nested job structures up to 7 levels deep.
        # It's verbose but avoids making separate API calls for each folder level.
        "tree":  "jobs[name,url,lastBuild[timestamp],jobs[name,url,lastBuild[timestamp],jobs[name,url,lastBuild[timestamp],jobs[name,url,lastBuild[timestamp],jobs[name,url,lastBuild[timestamp],jobs[name,url,lastBuild[timestamp],jobs[name,url,lastBuild[timestamp]]]]]]]]",
        "depth": 7
    }
    r = api_get(s, url, params=params, timeout=60)
    if r is None:
        return []

    data = r.json()
    jobs = []

    def flatten(job_list):
        """Recursively walk the job tree and collect leaf jobs (actual pipelines)."""
        for j in job_list:
            if "jobs" in j:
                # This is a folder — go deeper
                flatten(j["jobs"])
            else:
                # This is an actual job. Only include it if its last build
                # ran on or after FROM_DATE — no point fetching older jobs.
                last_build = j.get("lastBuild") or {}
                last_ts    = last_build.get("timestamp", 0)
                if last_ts >= cutoff_ms:
                    jobs.append(j)

    flatten(data.get("jobs", []))
    return jobs


def api_get(s, url, params=None, timeout=20, retries=3, delay=2):
    """
    A thin wrapper around requests.get() with retry logic.
    Jenkins can be slow or throw transient errors, especially when you're
    hitting it in parallel. Three retries with a short sleep handles most cases.
    Returns None if all retries fail so callers can handle it gracefully.
    """
    for attempt in range(retries):
        try:
            r = s.get(url, params=params, timeout=timeout)
            r.raise_for_status()
            return r
        except Exception:
            if attempt < retries - 1:
                time.sleep(delay)
    return None


def get_recent_builds(s, job_url, cutoff_ms, end_ms):
    """
    Fetch the last 100 builds for a job and filter down to those that
    started within our date range. We cap at 100 because that's usually
    more than enough for a day or two of builds.
    """
    url    = f"{job_url}api/json"
    params = {
        "tree":  "builds[number,timestamp,duration,result,displayName,url]{0,100}",
        "depth": 1
    }
    r = api_get(s, url, params=params)
    if r is None:
        return []
    builds = r.json().get("builds", [])
    # Filter to only builds that started within [cutoff_ms, end_ms)
    return [b for b in builds if cutoff_ms <= b.get("timestamp", 0) < end_ms]


def get_stages(s, build_url):
    """
    Fetch the stage breakdown for a specific build using the
    Jenkins Pipeline Stage View API (/wfapi/describe).
    This gives us each stage's name, status, and duration in ms.
    Returns an empty list if the build isn't a pipeline or the API fails.
    """
    url = f"{build_url}wfapi/describe"
    r   = api_get(s, url)
    if r is None:
        return []
    return r.json().get("stages", [])


def get_failure_reason(s, build_url):
    """
    For failed builds, try to pull a useful error message from the console log
    rather than just reporting 'FAILURE' with no context.

    Jenkins console logs can be huge, so we first check the total log size,
    then fetch only the last 8KB — that's almost always where the failure
    message ends up. We then scan backwards through the lines and match
    against common error patterns.

    Falls back to a generic 'See console log' message if nothing matches.
    """
    url = f"{build_url}logText/progressiveText"

    # First request just to get the total log size from the response header
    r = api_get(s, url, params={"start": 0})
    if r is None:
        return ""

    text_size = int(r.headers.get("X-Text-Size", 0))
    # Start from 8KB before the end to get the tail of the log
    start = max(0, text_size - 8000)

    r2 = api_get(s, url, params={"start": start})
    if r2 is None:
        return ""
    text = r2.text

    # These patterns cover most Jenkins failure scenarios — exceptions,
    # Groovy errors, plugin failures, non-zero exit codes, etc.
    error_patterns = [
        r"(ERROR:.*)",
        r"(FATAL:.*)",
        r"(Exception in thread.*)",
        r"(\w+Exception:.*)",
        r"(error:.*)",
        r"(Build step '.*' marked build as failure)",
        r"(script returned exit code \d+)",
        r"(hudson\..*Exception.*)",
        r"(Caused by:.*)",
        r"(Process leaked file descriptors.*)",
        r"(hudson\.plugins\..*)",
    ]

    # Scan from the bottom up — the last error is usually the relevant one
    lines = text.splitlines()
    for line in reversed(lines):
        line = line.strip()
        if not line:
            continue
        for pattern in error_patterns:
            m = re.search(pattern, line, re.IGNORECASE)
            if m:
                reason = m.group(1).strip()
                # Truncate long lines — 200 chars is enough for an Excel cell
                return reason[:200] if len(reason) > 200 else reason

    # Fallback: look for any line containing common failure keywords
    for line in reversed(lines):
        line = line.strip()
        if line and any(w in line.lower() for w in ["fail", "error", "abort", "exit"]):
            return line[:200]

    return "See console log"


def is_target_stage(stage_name):
    """
    Check whether a stage name matches our target stage keywords.
    Case-insensitive so "Fortify Scan", "FORTIFY SCAN", "fortify scan" all match.

    If you're adapting this script for a different stage, this is the function
    that decides what gets tracked — it just reads from STAGE_KEYWORDS at the top.
    """
    return any(kw in stage_name.lower() for kw in STAGE_KEYWORDS)


def ms_to_readable(ms):
    """
    Convert a millisecond duration into something human-readable.
    e.g. 754000ms → '12m 34s', 3661000ms → '1h 1m 1s'
    Returns an empty string for zero/negative values.
    """
    if ms <= 0:
        return ""
    s    = ms // 1000
    m, s = divmod(s, 60)
    h, m = divmod(m, 60)
    parts = []
    if h: parts.append(f"{h}h")
    if m: parts.append(f"{m}m")
    if s: parts.append(f"{s}s")
    return " ".join(parts) or "0s"


def fetch_build_data(s, job, cutoff_ms, end_ms):
    """
    The main per-job data collection function. For each build in the date range:
    - Gets the list of pipeline stages
    - Checks which stages match our target keyword
    - Pulls failure reason from the console log if the build failed
    - Builds a row dict for each matching stage (or one row if none found)

    This function is called in parallel by ThreadPoolExecutor, so keep it
    thread-safe — no shared mutable state, just return a list of dicts.
    """
    rows   = []
    builds = get_recent_builds(s, job["url"], cutoff_ms, end_ms)

    for build in builds:
        result = build.get("result") or "IN PROGRESS"
        stages = get_stages(s, build["url"])

        # Filter stages down to only the ones matching our target keywords
        target_stages = [st for st in stages if is_target_stage(st.get("name", ""))]

        # Convert the build timestamp to a readable local time.
        # UTC+3 is the timezone used here — adjust the hours=3 offset to match yours.
        build_time = datetime.fromtimestamp(
            build["timestamp"] / 1000, tz=timezone.utc
        ).astimezone(timezone(timedelta(hours=3)))

        # Only bother fetching the failure reason if the build actually failed
        failure_reason = ""
        if result == "FAILURE":
            failure_reason = get_failure_reason(s, build["url"])

        # Base row data — shared fields regardless of whether a target stage was found
        base = {
            "Job Name":              job.get("name", ""),
            "Job URL":               job.get("url", ""),
            "Build #":               build.get("number", ""),
            "Build URL":             build.get("url", ""),
            "Build Status":          result,
            "Failure Reason":        failure_reason,
            "Build Start":           build_time.strftime("%Y-%m-%d %H:%M:%S"),
            "Build Duration":        ms_to_readable(build.get("duration", 0)) or "0s",
            "Has Fortify Stage":     "Yes" if target_stages else "No",
            "Fortify Stage Name":    "",
            "Fortify Status":        "",
            "Fortify Duration":      "",
            "Fortify Duration (ms)": 0,
        }

        if target_stages:
            # A build could theoretically have multiple matching stages (e.g. if
            # Fortify runs in both pre-prod and prod stages). We create a separate
            # row for each so no data is lost.
            for fs in target_stages:
                row                          = base.copy()
                row["Fortify Stage Name"]    = fs.get("name", "")
                row["Fortify Status"]        = fs.get("status", "")
                row["Fortify Duration"]      = ms_to_readable(fs.get("durationMillis", 0)) or "0s"
                row["Fortify Duration (ms)"] = fs.get("durationMillis", 0)
                rows.append(row)
        else:
            # Build ran but didn't have a matching stage — still include it
            # in the 'All Builds' sheet so you have a complete picture
            rows.append(base)

    return rows


def collect_all_data(use_cache=False):
    """
    Orchestrates the full data collection:
    1. Fetch all jobs from Jenkins (filtered by date)
    2. For each job, fetch builds and stage data in parallel
    3. Cache the result to disk so reruns are fast

    Set use_cache=True if you've already fetched the data and just want to
    regenerate the Excel without hitting Jenkins again.
    """
    if use_cache and Path(CACHE_FILE).exists():
        print(f"[cache] Loading from {CACHE_FILE}")
        with open(CACHE_FILE) as f:
            return json.load(f)

    s                 = make_session()
    cutoff_ms, end_ms = get_date_range_ms()

    print("[1/3] Fetching job list (filtering by last build date)...")
    jobs = get_all_jobs(s, cutoff_ms)
    print(f"      Found {len(jobs)} jobs with builds since {FROM_DATE}")
    print(f"      Date range: {FROM_DATE} to {TO_DATE}")

    print("[2/3] Fetching builds & stage data (parallel)...")
    all_rows = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        # Submit one task per job — each task fetches builds + stages for that job
        futures = {
            pool.submit(fetch_build_data, s, job, cutoff_ms, end_ms): job
            for job in jobs
        }
        done = 0
        for future in as_completed(futures):
            done += 1
            rows = future.result()
            all_rows.extend(rows)
            print(f"      {done}/{len(jobs)} jobs processed, {len(all_rows)} rows so far", end="\r")

    print(f"\n      Total rows collected: {len(all_rows)}")

    # Save to cache so we can regenerate the Excel without re-fetching
    with open(CACHE_FILE, "w") as f:
        json.dump(all_rows, f, indent=2)
    print(f"[cache] Saved to {CACHE_FILE}")

    return all_rows


def make_doughnut(title, data_ref, label_ref, colors):
    """
    Create a doughnut chart for the Summary sheet.
    We use a doughnut (not pie) because it looks cleaner and the hole
    gives space to add a label if needed later.
    Colors are passed as hex strings to match each data category visually.
    """
    ch          = DoughnutChart()
    ch.title    = title
    ch.style    = 2
    ch.holeSize = 45
    ch.width    = 17
    ch.height   = 13
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(label_ref)
    series = ch.series[0]
    # Apply custom colours per slice — green for fast, amber for mid, red for slow
    for idx, hex_color in enumerate(colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = hex_color
        series.dPt.append(pt)
    return ch


def write_excel(rows, output_path):
    """
    Write all collected data to a formatted Excel workbook with three sheets:

    Sheet 1 — 'All Builds':
        Every build in the date range, with a Yes/No column indicating whether
        a target stage was detected. Useful for spotting pipelines that skipped
        the stage entirely.

    Sheet 2 — 'Fortify Stages Only':
        Filtered to only builds where the target stage was found. This is the
        main sheet most people will care about.

    Sheet 3 — 'Summary':
        High-level counts + a doughnut chart showing how scan durations are
        distributed across three bands: <10 mins, 10-15 mins, >15 mins.
        Adjust TEN_MIN_MS and FIFTEEN_MIN_MS below if your thresholds differ.
    """
    print("[3/3] Writing Excel report...")
    wb           = openpyxl.Workbook()
    ws_all       = wb.active
    ws_all.title = "All Builds"
    ws_fort      = wb.create_sheet("Fortify Stages Only")
    ws_sum       = wb.create_sheet("Summary")

    # Column definitions for each sheet
    all_columns = [
        "Job Name", "Build #", "Build Start", "Build Status", "Failure Reason",
        "Build Duration", "Has Fortify Stage", "Fortify Stage Name",
        "Fortify Status", "Fortify Duration", "Job URL", "Build URL"
    ]
    fort_columns = [
        "Job Name", "Build #", "Build Start", "Build Status", "Failure Reason",
        "Build Duration", "Fortify Stage Name", "Fortify Status",
        "Fortify Duration", "Job URL", "Build URL"
    ]

    # Styles — defined once and reused across both sheets
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # Dark navy header
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    YES_FILL    = PatternFill("solid", fgColor="C6EFCE")  # Green — Fortify stage present
    NO_FILL     = PatternFill("solid", fgColor="F2F2F2")  # Light grey — no Fortify stage
    FAIL_FILL   = PatternFill("solid", fgColor="FFC7CE")  # Red — failed build
    ALT_FILL    = PatternFill("solid", fgColor="EEF2FF")  # Alternating row shading
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Column widths — tuned so most content fits without manual resizing
    widths = {
        "Job Name": 38, "Build #": 10, "Build Start": 20,
        "Build Status": 14, "Failure Reason": 45,
        "Build Duration": 16, "Has Fortify Stage": 16,
        "Fortify Stage Name": 22, "Fortify Status": 14,
        "Fortify Duration": 16, "Job URL": 50, "Build URL": 50
    }

    # Columns that look better left-aligned (names, URLs, long text)
    LEFT_COLS = ("Job Name", "Build Start", "Fortify Stage Name",
                 "Failure Reason", "Job URL", "Build URL")

    def write_sheet(ws, data_rows, columns):
        """
        Write a header row + data rows to a worksheet with formatting.
        Adds auto-filter and freezes the header row so it stays visible
        when scrolling through large datasets.
        """
        ws.append(columns)
        for cell in ws[1]:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER

        ws.freeze_panes    = "A2"  # Keep header visible when scrolling
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

        for i, row in enumerate(data_rows, start=2):
            ws.append([row.get(c, "") for c in columns])
            alt = (i % 2 == 0)  # Alternating row colour for readability
            for j, col in enumerate(columns, start=1):
                cell           = ws.cell(row=i, column=j)
                cell.font      = Font(name="Arial", size=9)
                cell.alignment = LEFT if col in LEFT_COLS else CENTER

                # Conditional highlighting
                if col == "Has Fortify Stage":
                    cell.fill = YES_FILL if cell.value == "Yes" else NO_FILL
                elif col == "Build Status" and cell.value == "FAILURE":
                    cell.fill = FAIL_FILL
                elif alt:
                    cell.fill = ALT_FILL

        for j, col in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 15)

    write_sheet(ws_all, rows, all_columns)

    # Only pass builds that had the target stage to the second sheet
    fortify_rows = [r for r in rows if r.get("Has Fortify Stage") == "Yes"]
    write_sheet(ws_fort, fortify_rows, fort_columns)

    # Duration thresholds for the Summary chart — change these if your scans
    # are consistently faster or slower than these bands
    TEN_MIN_MS     = 10 * 60 * 1000   # 10 minutes in ms
    FIFTEEN_MIN_MS = 15 * 60 * 1000   # 15 minutes in ms

    # Count completed scans by duration band (only SUCCESS/UNSTABLE — skip failures)
    below_10  = sum(1 for r in fortify_rows if 0 < r["Fortify Duration (ms)"] < TEN_MIN_MS           and r["Build Status"] in ("SUCCESS", "UNSTABLE"))
    btw_10_15 = sum(1 for r in fortify_rows if TEN_MIN_MS <= r["Fortify Duration (ms)"] <= FIFTEEN_MIN_MS and r["Build Status"] in ("SUCCESS", "UNSTABLE"))
    above_15  = sum(1 for r in fortify_rows if r["Fortify Duration (ms)"] > FIFTEEN_MIN_MS            and r["Build Status"] in ("SUCCESS", "UNSTABLE"))

    # Deduplicate by (Build URL, Build #) to avoid counting multi-stage builds twice
    successful_fort_builds = len(set(
        (r["Build URL"], str(r["Build #"])) for r in fortify_rows
        if r["Build Status"] in ("SUCCESS", "UNSTABLE")
    ))
    total_successful_builds = len(set(
        (r["Build URL"], str(r["Build #"])) for r in rows
        if r["Build Status"] in ("SUCCESS", "UNSTABLE")
    ))

    # Summary metrics table
    summary = [
        ("From Date",                                 FROM_DATE),
        ("Total Successful Builds",                   total_successful_builds),
        ("Total Successful Builds with Fortify Scan", successful_fort_builds),
    ]

    ws_sum.column_dimensions["A"].width = 42
    ws_sum.column_dimensions["B"].width = 20

    ws_sum.append(["Metric", "Value"])
    for cell in ws_sum[1]:
        cell.font      = HEADER_FONT
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = CENTER

    for metric, value in summary:
        ws_sum.append([metric, value])

    for row in ws_sum.iter_rows(min_row=2):
        for cell in row:
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Chart data lives in cols G/H — far right so it doesn't clutter the summary table.
    # openpyxl needs the data in the sheet to build the chart reference.
    ws_sum["G1"] = "Duration Range"
    ws_sum["H1"] = "Scans"
    ws_sum["G1"].font = Font(bold=True, name="Arial", size=9, color="888888")
    ws_sum["H1"].font = Font(bold=True, name="Arial", size=9, color="888888")
    for i, (lbl, val) in enumerate([
        ("Below 10 mins",     below_10),
        ("10 to 15 mins",     btw_10_15),
        ("More than 15 mins", above_15),
    ], start=2):
        ws_sum[f"G{i}"] = lbl
        ws_sum[f"H{i}"] = val

    ws_sum.column_dimensions["G"].width = 22
    ws_sum.column_dimensions["H"].width = 10

    chart1 = make_doughnut(
        title     = f"Fortify Scan Duration Breakdown  ({FROM_DATE})",
        data_ref  = Reference(ws_sum, min_col=8, min_row=1, max_row=4),
        label_ref = Reference(ws_sum, min_col=7, min_row=2, max_row=4),
        colors    = ["1A7A4A", "E8A020", "C0392B"],  # Green / Amber / Red
    )
    chart1.width  = 15
    chart1.height = 14
    ws_sum.add_chart(chart1, "B6")

    wb.save(output_path)
    print(f"\n✅ Report saved → {output_path}")
    print(f"   All Builds sheet:     {len(rows)} rows")
    print(f"   Fortify Stages sheet: {len(fortify_rows)} rows")
    print(f"   Successful Fortify:   {successful_fort_builds}")
    print(f"   Chart — Duration:     <10m={below_10}, 10-15m={btw_10_15}, >15m={above_15}")


if __name__ == "__main__":
    # Set use_cache=True if you've already run a fetch and just want to
    # regenerate the Excel (useful while tweaking the report format)
    rows = collect_all_data(use_cache=False)
    write_excel(rows, get_output_filename())
